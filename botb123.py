from sqlalchemy import create_engine, Column, Integer, String, ForeignKey, BigInteger
from sqlalchemy.orm import sessionmaker, relationship, declarative_base
from telegram.ext import (Updater, CommandHandler, MessageHandler, Filters,
                          CallbackQueryHandler, ConversationHandler)
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from datetime import datetime, timedelta, date, time
from apscheduler.schedulers.background import BackgroundScheduler
import logging
import pytz
from pytz import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import random
from functools import partial

# Настройка логирования
logging.basicConfig(filename='myapp.log', level=logging.INFO,
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
logging.disable(logging.CRITICAL)
# Создание подключения к базе данных
engine = create_engine('mysql+pymysql://root:123@localhost/botdb',
                       pool_size=20,  # increases the number of persistent connections to 20
                       max_overflow=40  # allows up to 40 additional connections beyond the pool_size
                       )

Session = sessionmaker(bind=engine)
Base = declarative_base()


# Определение моделей
class Chat(Base):
    __tablename__ = 'chats'
    id = Column(BigInteger, primary_key=True)
    start_date = Column(String(255))


class ChatMember(Base):
    __tablename__ = 'members'
    id = Column(Integer, primary_key=True)
    chat_id = Column(BigInteger, ForeignKey('chats.id', ondelete='CASCADE'))
    user_id = Column(BigInteger, nullable=False)
    user_name = Column(String(255))
    full_name = Column(String(255))


class DailyRecord(Base):
    __tablename__ = 'daily_records'
    id = Column(Integer, primary_key=True)
    chat_member_id = Column(Integer, ForeignKey('members.id', ondelete='CASCADE'))
    date = Column(String(255))
    morning_hashtag = Column(String(255), default=False)
    evening_hashtag = Column(String(255), default=False)
    week_hashtag = Column(String(255), default=False)


class Settings(Base):
    __tablename__ = 'settings'
    id = Column(Integer, primary_key=True)
    chat_id = Column(BigInteger, unique=True, nullable=False)
    morning_hashtag = Column(String(255), default="#оу")
    evening_hashtag = Column(String(255), default="#ов")
    week_hashtag = Column(String(255), default="#неделя")
    morning_deadline = Column(String(255), default="10:00")
    evening_deadline = Column(String(255), default="23:59")
    start_date = Column(String(255), nullable=True)


class UserState(Base):
    __tablename__ = 'user_states'
    id = Column(Integer, primary_key=True)
    chat_id = Column(BigInteger, nullable=False)
    user_id = Column(BigInteger, nullable=False)
    state = Column(String(255))
    data = Column(String(255), nullable=True)


class Fine(Base):
    __tablename__ = 'fines'
    id = Column(Integer, primary_key=True)
    chat_member_id = Column(Integer, ForeignKey('members.id'))
    date_paid = Column(String(255))
    report_type = Column(String(255))
    fine_amount = Column(Integer, nullable=True)


Base.metadata.create_all(engine)


def get_settings(chat_id):
    session = Session()
    session.close()
    return session.query(Settings).filter_by(chat_id=chat_id).first() or Settings()


def is_admin(user_id, chat_id, bot):
    try:
        chat_administrators = bot.get_chat_administrators(chat_id)
        return any(admin.user.id == user_id for admin in chat_administrators)
    except:
        return False


# Функции для обработки хештегов и дедлайнов
def update_daily_record(chat_id, user_id, date, morning_hashtag=None, evening_hashtag=None, week_hashtag=None):
    session = Session()
    member = session.query(ChatMember).filter_by(chat_id=chat_id, user_id=user_id).first()

    if member:
        record = session.query(DailyRecord).filter_by(chat_member_id=member.id, date=date).first()
        if not record:
            # Создаем новую запись с утренним, вечерним и недельным хештегами
            record = DailyRecord(chat_member_id=member.id, date=date,
                                 morning_hashtag='1' if morning_hashtag else '0',
                                 evening_hashtag='1' if evening_hashtag else '0',
                                 week_hashtag='1' if week_hashtag else '0')
            session.add(record)
        else:
            # Обновляем только те хештеги, которые были переданы
            if morning_hashtag is not None:
                record.morning_hashtag = '1' if morning_hashtag else '0'
            if evening_hashtag is not None:
                record.evening_hashtag = '1' if evening_hashtag else '0'
            if week_hashtag is not None:
                record.week_hashtag = '1' if week_hashtag else '0'

        session.commit()
    session.close()


def create_user_mention(user_name, user_id, full_name):
    # Проверяем, действительно ли установлен user_name и он не пустой
    if user_name and user_name.strip() != "":
        return f"@{user_name}"
    else:
        # Если user_name отсутствует или пустой, используем user_id для создания упоминания
        return f"<a href='tg://user?id={user_id}'>{full_name}</a>"


def check_reports_and_notify(bot):
    logger.info("check_reports_and_notify: Начало функции")
    session = Session()
    current_time = datetime.now(pytz.timezone('Europe/Moscow'))
    today_date = current_time.date()
    is_weekday_sunday = current_time.weekday() == 6
    is_morning = current_time.time() < time(12, 0, 0)
    print(is_morning)

    logger.info(f"Function check_reports_and_notify started for {today_date}")

    chats = session.query(Chat).all()
    for chat in chats:
        logger.info(f"check_reports_and_notify: Обработка чата")
        chat_id = chat.id
        late_users_morning = []
        late_users_evening = []
        late_users_week = []

        logger.info(f"Processing chat: {chat_id}")
        try:
            logger.info("Начало обработки чата: {}".format(chat_id))
            members = session.query(ChatMember).filter_by(chat_id=chat_id).all()
            for member in members:
                logger.info(f"check_reports_and_notify: Обработка пользователя {member.user_id} в чате {chat_id}")
                record = session.query(DailyRecord).filter_by(chat_member_id=member.id, date=today_date).first()

                if is_morning and not is_weekday_sunday:
                    if not record or record.morning_hashtag == "0":
                        user_mention = create_user_mention(member.user_name, member.user_id, member.full_name)
                        late_users_morning.append(user_mention)

                if not is_morning and not is_weekday_sunday:
                    if not record or record.evening_hashtag == "0":
                        user_mention = create_user_mention(member.user_name, member.user_id, member.full_name)
                        late_users_evening.append(user_mention)

                if is_weekday_sunday and (not record or record.week_hashtag == "0"):
                    user_mention = create_user_mention(member.user_name, member.user_id, member.full_name)
                    late_users_week.append(user_mention)


            praise_messages = [
                "Превосходная работа! Ваши отчёты сияют как звёзды в ночном небе! 🌟",
                "Невероятный успех! Ваши отчеты отражают вашу страсть и упорство. 💪",
                "Браво! Каждый ваш отчёт - это шаг на пути к величию. 🚀",
                "Вы несомненно мастера своего дела! Ваши отчеты - пример для подражания. 👏",
                "Удивительно! Ваши отчеты свидетельствуют о вашем таланте и трудолюбии. 👌",
                "Вы превзошли сами себя! Ваши отчеты - это произведения искусства. 🎨",
                "Ваши отчеты как музыка для ушей, полные гармонии и мастерства. 🎶",
                "Каждый ваш отчёт - это очередной шедевр! Продолжайте в том же духе. 🌈",
                "Ваш труд не остался незамеченным. Ваши отчеты - пример высочайшего качества. 🏅",
                "Вы - звезда! Ваши отчеты озаряют путь к успеху. ✨",
                "Прекрасная работа! Ваши отчеты отражают вашу преданность и старание. 🌟",
                "Ваши отчеты - олицетворение профессионализма и внимания к деталям. 💼",
                "Так держать! Ваши отчеты каждый раз превосходят ожидания. 🚀",
                "Вы установили новый стандарт качества с вашими отчетами. Браво! 👍",
                "Каждый ваш отчёт - это путеводная звезда к мечтам и амбициям. 🌠",
                "Ваши отчеты - это праздник для глаз и ума. Великолепно! 🎉",
                "Вы вдохновляете нас всех! Ваши отчеты - пример настойчивости и целеустремленности. 💖",
                "Ваше усердие в отчетах заслуживает самых высоких похвал. Впечатляюще! 👏",
                "Ваши отчеты - как свежий ветер, приносящий новые идеи и перспективы. 🌬️",
                "Вы сияете ярче всех! Ваши отчеты - это воплощение совершенства. 🌟",
                "Ваши отчеты озаряют путь к успеху, как солнце освещает утро. ☀️",
                "Такое мастерство! Ваши отчеты - пример уникальности и креативности. 🎨",
                "Вы - истинные вдохновители! Ваши отчеты показывают, на что вы способны. 🌈",
                "Ваши отчеты - это мост в будущее полное достижений и успехов. 🌉",
                "Вы доказали, что ничего невозможного нет. Ваши отчеты - это подтверждение. 💪",
                "Ваши отчеты - это песня успеха, наполненная гармонией и мелодией. 🎵",
                "Ваши отчеты - это легенда о вашем труде и старании. Легендарно! 🏆",
                "Такое внимание к деталям! Ваши отчеты - пример исключительной работы. 🔍",
                "Вы - источник вдохновения! Ваши отчеты - это воплощение вашего духа. 💫",
                "Ваши отчеты - как маяк, освещающий путь к цели. Блестяще! 🚩",
                "Ваши отчеты - это путешествие в мир качества и совершенства. 🗺️",
                "Ваши отчеты - это поэзия успеха, написанная вашими руками. 📜",
                "Ваши отчеты - это мозаика успеха, собранная из мелких деталей мастерства. 🎭",
                "Вы - чемпион в создании отчетов, каждый из которых - триумф усердия! 🏆",
                "Ваши отчеты - как музыкальная симфония, в которой каждая нота на своем месте. 🎼",
                "Вы разгадали код успеха! Ваши отчеты - это шифр совершенства. 🗝️",
                "Каждый ваш отчет - как капля в океане ваших достижений. Невероятно! 💧",
                "Ваши отчеты - это северное сияние в мире работы. Завораживающе! 🌌",
                "Вы воплотили ваши мечты в каждом отчете. Ваши усилия не зря! 💭",
                "Вы взлетели к вершинам мастерства в ваших отчетах. Потрясающе! 🚀",
                "Ваши отчеты - это оазис в пустыне повседневности. Восхитительно! 🌴",
                "Вы пишете историю успеха с каждым отчетом. Вдохновляюще! 📚",
                "Ваши отчеты - как зеркало вашей души, отражающее ваши усилия. 🗺️",
                "Каждый ваш отчет - это кирпичик в стене вашего профессионального роста. 🧱",
                "Ваши отчеты - как вспышка света в темноте, освещающая путь к успеху. 💡",
                "Вы - маэстро в мире отчетов. Ваше мастерство поражает! 🎻",
                "Каждый ваш отчет - это отголосок вашего неутомимого стремления к совершенству. 🌟",
                "Вы - алхимики успеха, превращающие каждый отчет в золото. 🧪",
                "Ваши отчеты - это танец слов и цифр, создающий гармонию успеха. 💃",
                "Ваши отчеты - это гимн вашему упорству и таланту. Непревзойденно! 🎵",
                "Вы рисуете картину своего успеха через каждый отчет. Вдохновляете! 🖼️",
                "Ваши отчеты - как кулинарное произведение, приготовленное с любовью и мастерством. 🍳",
                "Каждый ваш отчет - это шаг к вершине ваших возможностей. Восхищаюсь вами! 🏔️",
                "Ваши отчеты - как рассвет нового дня, полного новых возможностей. 🌅",
                "Вы - мастер слов и анализа. Ваши отчеты - это ваше искусство. 📖",
                "Каждый ваш отчет - это волна инноваций и прогресса. Отлично! 🌊",
                "Ваши отчеты - это путеводный свет в мире постоянного развития. 🔦",
                "Вы - архитектор вашего успеха, а ваши отчеты - это его фундамент. 🏗️",
                "Ваши отчеты - это сад ваших достижений, где каждый цветок - это ваш труд. 🌺",
                "Вы вдыхаете жизнь в каждый отчет, делая его живым и динамичным. 🍃",
                "Ваши отчеты - как радуга после дождя, полная надежды и света. 🌈",
                "Вы - рыцари на полях отчетности, сражающиеся за качество и точность. 🛡️",
                "Каждый ваш отчет - это звездопад ваших достижений, освещающий путь другим. ✨"
            ]
            random_praise = random.choice(praise_messages)

            if not is_weekday_sunday:
                if is_morning and late_users_morning:
                    send_notification(bot, chat_id, late_users_morning, "утренний отчёт")
                    logger.info(f"check_reports_and_notify: Отправка уведомлений о пропущенных утренних отчетах в чат {chat_id}")
                else:
                    logger.info(f"No late users for morning report in chat {chat_id}")
                if late_users_evening and not is_morning:
                    send_notification(bot, chat_id, late_users_evening, "вечерний отчёт")
                    logger.info(f"check_reports_and_notify: Отправка уведомлений о пропущенных вечерних отчетах в чат {chat_id}")
                else:
                    logger.info(f"No late users for evening report in chat {chat_id}")

            if is_weekday_sunday and late_users_week:
                send_notification(bot, chat_id, late_users_week, "недельный отчёт")
                logger.info(f"check_reports_and_notify: Отправка уведомлений о пропущенных недельных отчетах в чат {chat_id}")
            else:
                logger.info(f"No late users for weekly report in chat {chat_id}")

            if not is_weekday_sunday:
                if is_morning and not late_users_morning:
                    bot.send_message(chat_id=chat_id,
                                     text="Все участники сдали утренние отчеты вовремя. Молодцы! " + random_praise)
                    logger.info(f"All morning reports submitted on time in chat {chat_id}")
                if not is_morning and not late_users_evening:
                    bot.send_message(chat_id=chat_id,
                                     text="Все участники сдали вечерние отчеты вовремя. Молодцы! " + random_praise)
                    logger.info(f"All evening reports submitted on time in chat {chat_id}")
            if is_weekday_sunday:
                if not late_users_week:
                    bot.send_message(chat_id=chat_id,
                                     text="Все участники сдали недельные отчеты вовремя. Отличная работа! " + random_praise)
                    logger.info(f"All weekly reports submitted on time in chat {chat_id}")
            logger.info("Завершение обработки чата: {}".format(chat_id))
        except Exception as e:
            logger.error("Ошибка в чате {}: {}".format(chat_id, str(e)))

    session.close()
    logger.info("Function check_reports_and_notify completed")


def send_notification(bot, chat_id, user_list, report_type):
    session = Session()
    chat = session.query(Chat).filter(Chat.id == chat_id).first()
    if chat and user_list:
        current_date = datetime.now().date()
        start_date = datetime.strptime(chat.start_date, "%Y-%m-%d").date()
        days_since_start = (current_date - start_date).days

        if days_since_start >= 5:
            late_message_variants = [
                "Вы сегодня, наверняка, были так заняты покорением мира своей красотой, что забыли сдать  "
                "отчёт. Cоизвольте внести так же свой вклад в помощь нуждающимся в Вашем внимании 💖",
                "Ваша занятость сегодня, без сомнения, была направлена на великие дела, но не забывайте о маленьких "
                "победах в виде сданных отчётов 🌟",
                "Похоже, сегодня у вас был день полный чудес и волшебства, но не забывайте добавить к ним отчёт ✨",
                "Ваши грандиозные планы на сегодня, видимо, отняли все время. Но не упускайте момент поделиться с нами "
                "своими "
                "успехами в отчёте 🚀",
                "Ваше мастерство управления временем сегодня, кажется, подвело. Не забывайте найти минутку для отчёта 😊",
                "Сегодня вы, наверное, творили чудеса, но не забывайте о магии отчёта. Может, ваши чудеса принесут "
                "радость "
                "не только вам, но и кому-то еще? 🌟✨",
                "Вы, как волшебник, сегодня были заняты созданием чудес, но помните: каждый отчёт – это еще одно "
                "волшебство. И не забудьте о маленьком вкладе в большое дело! 🎩🌈",
                "Ваши героические свершения сегодня, вероятно, не оставили времени для отчёта. Но помните: каждый "
                "вклад в "
                "отчёт - это шанс сделать мир лучше! 🚀💫",
                "Похоже, сегодня вы были заняты покорением мира, но не забывайте о маленьких шагах, таких как отчёт, "
                "и о маленьких жестах, которые могут значить много! 🌍❤️",
                "Ваша занятость сегодня, без сомнения, была великолепна, но не забывайте о волшебстве маленьких вещей, "
                "как сданные отчёты. Ведь они тоже помогают делать мир лучше! 🌟",
                "Кажется, ваш день был наполнен чудесами, но волшебный мир отчётов тоже ждал вашего вклада. Подарите "
                "им "
                "частичку вашего внимания! ✨",
                "Ваши грандиозные планы на сегодня, видимо, забрали все время. Не пропустите шанс оставить свой след в "
                "благотворительном вкладе! 🚀",
                "Сегодняшний день, вероятно, был полон событий, но не забудьте о маленьком, но важном вкладе в наш "
                "общий "
                "мир. 😊",
                "Ваше мастерство сегодня зашкаливает, но не забывайте об отчёте и поделитесь этим волшебством в нашем "
                "общем сундучке "
                "добрых дел! 🌈",
                "Сегодня вы, должно быть, спасали мир! Не забудьте ещё и маленькое чудо сделать. 🌍",
                "Ваш день был наверняка полон героических свершений! Не упустите момент добавить к ним ещё одно — ваш "
                "вклад в нашу общую историю. 🦸‍♂",
                "Вы, как всегда, блестяще справляетесь с задачами! Но не забывайте рассказать о них в отчёте и "
                "оставить свой след в фонде добрых дел. "
                "💫",
                "Вы сегодня, как волшебник, творили чудеса! Не упустите шанс добавить волшебный штрих в коллекцию "
                "добрых "
                "дел. 🧙‍♂️",
                "Ваш день был, наверное, полон приключений! Не забудьте оставить след вашего приключения в казне "
                "великолепных свершений. 🚀",
                "Вы сегодня были, как супергерой, везде и всюду! Поделитесь частицей вашей суперсилы с командой добрых "
                "дел. 🦸‍♀️",
                "Сегодня вы, наверняка, творили историю! Поделитесь этой историей с фондом добрых дел. 📜",
                "Ваш день был, как эпос! Добавьте ещё одну главу, внесите свой вклад в книгу добрых дел. 📖",
                "Вы сегодня были звездой на небосводе! Не забудьте оставить искру в созвездии добрых дел. ⭐",
                "Ваша энергия сегодня, наверное, освещала города! Поделитесь этим светом, внесите свой вклад в наш "
                "фонд "
                "светлых начинаний. 💡",
                "Вы сегодня, кажется, вдохновляли мир! Не упустите шанс вдохновить и нас, добавьте свой вклад в фонд "
                "светлых свершений. 🌠",
                "Сегодня вы, наверное, покоряли вершины! Не забудьте поделиться этим успехом, добавьте свой вклад в "
                "фонд "
                "светлых свершений.. 🏔️",
                "Ваш день был, как карнавал! Поделитесь этим праздником, внесите свой вклад в фонд радостных "
                "моментов. 🎉",
                "Вы, наверняка, сегодня разгадывали тайны вселенной! Не упустите момент поделиться этими знаниями, "
                "добавив свой вклад в нашу вселенную добра. 🌌",
                "Ваш день был полон магии! Продолжайте творить чудеса, внося свой вклад в фонд волшебства. ✨",
                "Кажется, сегодня вы летали выше облаков! Не забудьте поделиться этим ощущением, внесите свой вклад в "
                "фонд светлых свершений. 🌤️",
                "Вы сегодня, вероятно, писали шедевры! Продолжайте творить искусство, добавив свой вклад в нашу "
                "галерею "
                "добрых дел. 🎨",
                "Ваш день был как великая одиссея! Не забудьте добавить главу об этом в нашу книгу великих и добрых "
                "дел. 📚",
                "Сегодня вы, наверное, укрощали бури! Поделитесь этой силой, внесите свой вклад в фонд "
                "благотворительности. 🌪️",
                "Вы сегодня, кажется, раскрашивали мир! Не упустите шанс раскрасить и картину добрых дел. 🖌️",
                "Сегодня вы, вероятно, творили чудеса! Не забудьте добавить свою волшебную пыльцу в фонд чудес. 🌟",
                "Ваш день был как грандиозное представление! Продолжайте удивлять, добавив свой вклад в фонд светлых "
                "свершений.. 🎭",
                "Кажется, сегодня вы собирали звёзды с небес! Поделитесь этим сиянием, добавьте искорку в фонд светлых "
                "свершений.. ✨",
                "Сегодня вы, наверняка, вели танец судьбы! Продолжайте вдохновлять, добавив свой шаг в наш танец "
                "добрых "
                "дел. 💃",
                "Ваш день был как невероятное путешествие! Оставьте свой след в путеводителе великих историй. 🗺️",
                "Сегодня вы, наверняка, плавали по волнам вдохновения! Продолжайте плыть, добавив свой вклад в фонд "
                "добрых дел. 🌊",
                "Ваш день был полон сверкающих моментов! Заставьте светить и наш мир, добавив свой вклад в наш фонд "
                "добра "
                "и благотворительности. 💎",
                "Сегодня вы, наверняка, творили музыку жизни! Продолжайте компонировать, добавив свою ноту в нашу "
                "симфонию добра. 🎶",
                "Вы сегодня, вероятно, путешествовали по облакам! Поделитесь этими высотами, добавив свой вклад в наш "
                "фонд настоящей доброты. ☁ "
            ]
            additional_text = random.choice(late_message_variants) + " Отправьте на любую благотворительность 250₽, " \
                                                                     "и пришлите сюда в чат скриншот перевода!😉 " \
                                                                     "Согласно принятым всеми вами правилам, " \
                                                                     "за опоздание, даже минутное, мы помогаем " \
                                                                     "другим 😊🌸 "
        else:
            # Вариант сообщения в первые 4 дня курса
            additional_text = " Пожалуйста, не забудьте сдать его в ближайшее время! 😊"
        message_text = f"{report_type} не отправили вовремя: " + ", ".join(user_list) + ". " + additional_text
        bot.send_message(chat_id=chat_id, text=message_text, parse_mode="HTML")

    session.close()


# Функции для обработки команд
def start(update, context):
    update.message.reply_text('Привет! Я бот для отслеживания хештегов курса.')


def show_buttons(update, context):
    keyboard = [
        [InlineKeyboardButton("Отправить сводку в ЛС", callback_data='send_report_in_private')],
        [InlineKeyboardButton("Удалить участника", callback_data='show_participants')],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    update.message.reply_text('Выберите действие:', reply_markup=reply_markup)


def button(update, context):
    query = update.callback_query
    query.answer()
    chat_id = query.message.chat_id
    user_id = query.from_user.id

    callback_data = query.data
    bot = context.bot

    # Проверка, является ли пользователь администратором
    if not is_admin(user_id, chat_id, bot):
        query.edit_message_text(text="Только администраторы могут использовать эту команду.")
        return

        # Обработка действий в зависимости от callback_data
    if callback_data.startswith('remove_'):
        user_id_to_remove = int(callback_data.split('_')[1])
        remove_member_from_chat(chat_id, user_id_to_remove)
        query.edit_message_text(text=f"Участник с ID {user_id_to_remove} удален.")
    elif callback_data == 'show_participants':
        show_participants(update, context)
    elif callback_data == 'send_report_in_private':
        try:
            send_excel_file_in_private(update, context)
            query.edit_message_text(text="Отчёт был отправлен вам в ЛС.")
        except Exception as e:
            query.edit_message_text(text="Ошибка при отправке отчёта: " + str(e))
        return

    logger.info(f"Нажатие кнопки: {callback_data} пользователем {user_id} в чате {chat_id}")


def join(update, context):
    chat_id = update.message.chat_id
    user_id = update.message.from_user.id
    user_name = update.message.from_user.username
    first_name = update.message.from_user.first_name
    last_name = update.message.from_user.last_name


    try:
        response = add_member_to_chat(chat_id, user_id, user_name, first_name, last_name)
        update.message.reply_text(response)
    except Exception as e:
        update.message.reply_text(f'Ошибка при добавлении участника: {e}')


def add_member_to_chat(chat_id, user_id, user_name, first_name, last_name):
    session = Session()
    full_name = f"{first_name or ''} {last_name or ''}".strip()
    member = session.query(ChatMember).filter_by(chat_id=chat_id, user_id=user_id).first()

    if not member:
        # Логгирование добавления нового участника
        logger.info(f"Adding new member: User ID {user_id}, Chat ID {chat_id}, Username {user_name}")
        member = ChatMember(chat_id=chat_id, user_id=user_id, user_name=user_name, full_name=full_name)
        session.add(member)
        session.commit()
        session.close()
        logger.info("New member added successfully.")
        return f'Вы успешно добавлены в список участников!'
    else:
        # Логгирование обновления существующего участника
        logger.info(f"Member already exists. Checking for updates: User ID {user_id}, Chat ID {chat_id}")
        if member.user_name != user_name or member.full_name != full_name:
            logger.info(f"Updating member info: Old Username {member.user_name}, New Username {user_name}")
            member.user_name = user_name
            member.full_name = full_name
            session.commit()
            logger.info("Member info updated successfully.")
        session.close()
        return 'Вы уже зарегистрированы.'


def button_callback_handler(update, context):
    query = update.callback_query
    query.answer()

    # Предполагаем, что в `callback_data` содержится информация для добавления участника
    user_id_to_add = query.data
    # Здесь должна быть логика для добавления участника по user_id
    query.edit_message_text(text=f"Участник с ID {user_id_to_add} добавлен.")


def add_member(user_id, chat_id):
    session = Session()
    # Проверяем, существует ли уже участник в базе данных
    existing_member = session.query(ChatMember).filter_by(user_id=user_id, chat_id=chat_id).first()
    if not existing_member:
        # Добавляем нового участника
        new_member = ChatMember(user_id=user_id, chat_id=chat_id)
        session.add(new_member)
        session.commit()
        session.close()
        return f"Участник с ID {user_id} добавлен."
    else:
        session.close()
        return f"Участник с ID {user_id} уже существует."


def create_member_buttons(bot, chat_id):
    chat_members = bot.get_chat_administrators(chat_id)
    keyboard = []

    for member in chat_members:
        # Исключаем администратора (бота) из списка кнопок
        if member.user.id != bot.id:
            button_text = f"{member.user.first_name} {member.user.last_name or ''}"
            callback_data = f"add_{member.user.id}"
            keyboard.append([InlineKeyboardButton(button_text, callback_data=callback_data)])

    return InlineKeyboardMarkup(keyboard)


def send_excel_file_in_private(update, context):
    # Получение chat_id и user_id из запроса
    query = update.callback_query
    user_id = query.from_user.id

    # Проверка существования сообщения от пользователя
    if query.message:
        chat_id = query.message.chat_id
    else:
        # Если сообщение отсутствует, используем user_id в качестве chat_id
        chat_id = user_id

    # Вызов функции создания Excel-файла
    file_path = create_excel_file(chat_id)

    try:
        with open(file_path, 'rb') as file:
            context.bot.send_document(chat_id=user_id, document=file)
    except Exception as e:
        query.edit_message_text(text="Ошибка при отправке отчёта: " + str(e))


def send_final_reminder(bot, report_type):
    session = Session()
    current_time = datetime.now(pytz.timezone('Europe/Moscow'))
    today_date = current_time.date()

    # Формируем условие фильтрации в зависимости от типа отчета
    if report_type == "утреннего":
        condition = (DailyRecord.morning_hashtag == 'False') | (DailyRecord.morning_hashtag == None)
    elif report_type == "вечернего":
        condition = (DailyRecord.evening_hashtag == 'False') | (DailyRecord.evening_hashtag == None)
    elif report_type == "недельного" and current_time.weekday() == 6:
        condition = (DailyRecord.week_hashtag == 'False') | (DailyRecord.week_hashtag == None)
    else:
        # Если тип отчета не подходит под критерии, прекращаем выполнение функции
        return

    # Получаем все записи DailyRecord и соответствующих ChatMember за сегодня
    records = session.query(ChatMember, DailyRecord).join(DailyRecord,
                                                          ChatMember.id == DailyRecord.chat_member_id).filter(
        DailyRecord.date == today_date, condition).all()

    # Словарь для хранения пользователей, опаздывающих с отчетами, по чатам
    late_users_by_chat = {}

    for member, record in records:
        late_users_by_chat.setdefault(member.chat_id, []).append(
            create_user_mention(member.user_name, member.user_id, member.full_name))

    # Отправка уведомлений для каждого чата
    for chat_id, users in late_users_by_chat.items():
        if users:
            bot.send_message(chat_id=chat_id,
                             text=f"Напоминание: осталось 15 минут на сдачу {report_type} отчёта. Не отправили отчёт: " + ", ".join(
                                 users))

    session.close()


def check_hashtags_and_notify(bot):
    def job_function():
        session = Session()
        logger.info("check_hashtags_and_notify: Начало функции")
        # Установка временной зоны в Московское время
        moscow_tz = pytz.timezone('Europe/Moscow')
        current_time = datetime.now(moscow_tz)
        today_date = current_time.date()

        morning_end_time = time(10, 0, 0)
        evening_end_time = time(23, 59, 59)

        chats_to_notify = session.query(Chat).all()
        for chat in chats_to_notify:
            logger.info(f"Checking chat: {chat.id}")
            morning_late_users = []
            evening_late_users = []
            week_late_users = []
            all_morning_reports_submitted = True
            is_weekday_sunday = current_time.weekday() == 6

            members = session.query(ChatMember).filter_by(chat_id=chat.id).all()
            for member in members:
                logger.info(f"check_hashtags_and_notify: Обработка пользователя {member.user_id}")
                record = session.query(DailyRecord).filter_by(chat_member_id=member.id, date=today_date).first()

                if not record:
                    record = DailyRecord(chat_member_id=member.id, date=today_date.strftime('%Y-%m-%d'),
                                         morning_hashtag=False, evening_hashtag=False, week_hashtag=False)
                    session.add(record)

                user_mention = create_user_mention(member.user_name, member.user_id, member.full_name)

                if current_time.time() >= morning_end_time and not record.morning_hashtag:
                    morning_late_users.append(user_mention)
                    all_morning_reports_submitted = False
                    record.morning_hashtag = False

                if current_time.time() >= evening_end_time and not record.evening_hashtag:
                    evening_late_users.append(user_mention)
                    record.evening_hashtag = False

                if current_time.weekday() == 6 and (not record or not record.week_hashtag):
                    week_late_users.append(user_mention)
                    if record:
                        record.week_hashtag = False

            session.commit()
        session.close()
        logger.info("check_hashtags_and_notify: Конец функции")

    return job_function


def send_hour_reminder(bot, chat_id, report_type):
    bot.send_message(chat_id=chat_id,
                     text=f"Напоминание: остался 1 час на сдачу {report_type} отчёта. Пожалуйста, убедитесь, что вы отправили ваш отчёт.")


def send_fifteen_minute_reminder(bot, chat_id, report_type):
    try:
        session = Session()
        current_time = datetime.now(pytz.timezone('Europe/Moscow'))
        today_date = current_time.date()
        late_users = []
        members = session.query(ChatMember).filter_by(chat_id=chat_id).all()

        for member in members:
            record = session.query(DailyRecord).filter_by(chat_member_id=member.id, date=today_date).first()
            if not record or (report_type == "утреннего" and record.morning_hashtag == "0") or (
                    report_type == "вечернего" and record.evening_hashtag == "0"):
                user_mention = create_user_mention(member.user_name, member.user_id, member.full_name)
                late_users.append(user_mention)

        if late_users:
            message_text = f"Напоминание: осталось 15 минут на сдачу {report_type} отчёта. Не отправили отчёт: " + ", ".join(
                late_users)
            bot.send_message(chat_id=chat_id, text=message_text, parse_mode="HTML")
        else:
            logger.info(f"No late users for {report_type} report in chat {chat_id}")

        session.close()
    except Exception as e:
        session.close()
        logger.error(f"Error in send_fifteen_minute_reminder: {e}")


def reschedule_jobs(scheduler, bot):
    try:
        session = Session()
        chat_list = session.query(Chat).all()

        moscow_tz = pytz.timezone('Europe/Moscow')

        for chat in chat_list:
            chat_id = chat.id

            # Фиксированное время напоминаний
            morning_time = time(10, 0)  # Утренний дедлайн в 10:00
            evening_time = time(23, 59)  # Вечерний дедлайн в 23:59
            week_reminder_time = time(23, 59)  # Недельный дедлайн в 17:00 в воскресенье

            # Уникальные идентификаторы заданий
            morning_hour_reminder_id = f"morning_hour_reminder_{chat_id}"
            morning_fifteen_minute_reminder_id = f"morning_15min_reminder_{chat_id}"
            evening_hour_reminder_id = f"evening_hour_reminder_{chat_id}"
            evening_fifteen_minute_reminder_id = f"evening_15min_reminder_{chat_id}"
            week_fifteen_minute_reminder_id = f"week_15min_reminder_{chat_id}"
            # Идентификаторы для новых задач
            reports_notify_morning_id = f"reports_notify_morning_{chat_id}"
            reports_notify_evening_id = f"reports_notify_evening_{chat_id}"
            reports_notify_week_id = f"reports_notify_week_{chat_id}"
            morning_job_id = f"hashtags_notify_morning_{chat_id}"
            evening_job_id = f"hashtags_notify_evening_{chat_id}"
            week_job_id = f"hashtags_notify_week_{chat_id}"

            # Рассчет времени для напоминаний
            morning_hour_reminder = (datetime.combine(datetime.today(), morning_time) - timedelta(hours=1)).time()
            morning_fifteen_minute_reminder = (
                    datetime.combine(datetime.today(), morning_time) - timedelta(minutes=15)).time()
            evening_hour_reminder = (datetime.combine(datetime.today(), evening_time) - timedelta(hours=1)).time()
            evening_fifteen_minute_reminder = (
                    datetime.combine(datetime.today(), evening_time) - timedelta(minutes=15)).time()
            # week_hour_reminder = (datetime.combine(datetime.today(), week_reminder_time) - timedelta(hours=1)).time()
            week_fifteen_minute_reminder = (
                    datetime.combine(datetime.today(), week_reminder_time) - timedelta(minutes=15)).time()

                # Добавление задач напоминания в планировщик
            if not scheduler.get_job(morning_hour_reminder_id):
                scheduler.add_job(send_hour_reminder, 'cron', hour=morning_hour_reminder.hour,
                                      minute=morning_hour_reminder.minute, args=(bot, chat_id, "утреннего"),
                                  day_of_week='mon,tue,wed,thu,fri,sat',
                                      timezone=moscow_tz)
            if not scheduler.get_job(morning_fifteen_minute_reminder_id):
                scheduler.add_job(send_fifteen_minute_reminder, 'cron', hour=morning_fifteen_minute_reminder.hour,
                                      minute=morning_fifteen_minute_reminder.minute, args=(bot, chat_id, "утреннего"),
                                  day_of_week='mon,tue,wed,thu,fri,sat',
                                      timezone=moscow_tz)
            if not scheduler.get_job(evening_hour_reminder_id):
                scheduler.add_job(send_hour_reminder, 'cron', hour=evening_hour_reminder.hour,
                                      minute=evening_hour_reminder.minute, args=(bot, chat_id, "вечернего"),
                                  day_of_week='mon,tue,wed,thu,fri,sat',
                                      timezone=moscow_tz)
            if not scheduler.get_job(evening_fifteen_minute_reminder_id):
                scheduler.add_job(send_fifteen_minute_reminder, 'cron', hour=evening_fifteen_minute_reminder.hour,
                                      minute=evening_fifteen_minute_reminder.minute, args=(bot, chat_id, "вечернего"),
                                  day_of_week='mon,tue,wed,thu,fri,sat',
                                      timezone=moscow_tz)
            if not scheduler.get_job(week_fifteen_minute_reminder_id):
                scheduler.add_job(send_fifteen_minute_reminder, 'cron', day_of_week='sun',
                                      hour=week_fifteen_minute_reminder.hour, minute=week_fifteen_minute_reminder.minute,
                                      args=(bot, chat_id, "недельного"), timezone=moscow_tz)

                # Добавление новых задач, если они еще не существуют
            # job_function = check_hashtags_and_notify(bot)
        if not scheduler.get_job(morning_job_id):
            scheduler.add_job(check_hashtags_and_notify, 'cron', id=morning_job_id,
                                  day_of_week='mon,tue,wed,thu,fri,sat', hour=10, minute=1, args=[bot],
                                  timezone=moscow_tz)

        if not scheduler.get_job(evening_job_id):
            scheduler.add_job(check_hashtags_and_notify(bot), 'cron', id=evening_job_id,
                                  day_of_week='mon,tue,wed,thu,fri,sat', hour=23, minute=59, second=59,
                                  timezone=moscow_tz)

        if not scheduler.get_job(week_job_id):
            scheduler.add_job(check_hashtags_and_notify(bot), 'cron', id=week_job_id, day_of_week='sun', hour=23, minute=59,
                                  timezone=moscow_tz)

        if not scheduler.get_job(reports_notify_morning_id):
            scheduler.add_job(check_reports_and_notify, 'cron', id=reports_notify_morning_id,
                                  day_of_week='mon,tue,wed,thu,fri,sat', hour=10, minute=1, args=[bot],
                                  timezone=moscow_tz)

        if not scheduler.get_job(reports_notify_evening_id):
            scheduler.add_job(lambda: check_reports_and_notify(bot), 'cron', id=reports_notify_evening_id,
                                      day_of_week='mon,tue,wed,thu,fri,sat', hour=23, minute=59, second=59,
                                      timezone=moscow_tz)

        if not scheduler.get_job(reports_notify_week_id):
            scheduler.add_job(lambda: check_reports_and_notify(bot), 'cron', id=reports_notify_week_id,
                                      day_of_week='sun', hour=23, minute=59, second=59, timezone=moscow_tz)

        session.close()
        logger.info("reschedule_jobs: Задачи успешно добавлены")
    except Exception as e:
        session.close()
        logger.error(f"reschedule_jobs: Ошибка при добавлении задач: {e}")


def is_valid_week_report(date_str):
    """Проверяет, является ли дата воскресеньем."""
    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
    return date_obj.weekday() == 6  # Воскресенье


def cancel(update, context):
    update.message.reply_text('Операция отменена.')
    return ConversationHandler.END


def sanitize_sheet_title(title):
    invalid_chars = ['*', '/', '\\', '?', ':', '[', ']']
    for char in invalid_chars:
        title = title.replace(char, '')  # Заменяем запрещенные символы на пустую строку
    return title[:31]  # Обрезаем название до 31 символа, если оно слишком длинное


def create_excel_file(chat_id):
    session = Session()

    members = session.query(ChatMember).filter(ChatMember.chat_id == chat_id).order_by(ChatMember.full_name).all()
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for member in members:
        raw_title = (member.full_name or member.user_name or str(member.user_id))[:31]
        sheet_title = sanitize_sheet_title(raw_title)
        ws = wb.create_sheet(title=sheet_title)

        headers = ["Дата", "Утренний отчёт", "Вечерний отчёт", "Недельный отчёт", "Штраф", "Дата уплаты штрафа",
                   "За отчёт"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.border = thin_border

        records = session.query(DailyRecord).filter_by(chat_member_id=member.id).order_by(DailyRecord.date.desc()).all()
        fines = session.query(Fine).filter_by(chat_member_id=member.id).order_by(Fine.date_paid).all()

        for record in records:
            date_obj = datetime.strptime(record.date, '%Y-%m-%d').date()
            is_weekday_sunday = date_obj.weekday() == 6  # Проверяем, воскресенье ли это

            morning_report = "" if is_weekday_sunday else record.morning_hashtag
            evening_report = "" if is_weekday_sunday else record.evening_hashtag
            week_report = record.week_hashtag if is_weekday_sunday else ""

            row = [record.date, morning_report, evening_report, week_report, "", "", ""]
            ws.append(row)

            for cell in ws[ws.max_row]:
                cell.border = thin_border
                if cell.column_letter in ['B', 'C', 'D']:
                    if cell.value == '1':
                        cell.fill = green_fill
                    elif cell.value == '0':
                        cell.fill = red_fill
                    elif cell.value == 'fine':
                        cell.fill = yellow_fill

        for fine in fines:
            fine_row = ["", "", "", "", "Штраф", fine.date_paid, fine.report_type]
            ws.append(fine_row)
            for cell in ws[ws.max_row]:
                cell.border = thin_border
                cell.fill = yellow_fill

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 20

    file_path = 'report.xlsx'
    wb.save(file_path)
    session.close()
    return file_path


def send_excel_file(update, context):
    chat_id = update.message.chat_id
    file_path = create_excel_file(chat_id)
    with open(file_path, 'rb') as file:
        context.bot.send_document(chat_id=chat_id, document=file)


def add_member_to_chat(chat_id, user_id, user_name, first_name, last_name):
    session = Session()
    # Создание полного имени, заменяя отсутствующие значения на пробел
    full_name = f"{first_name or ''} {last_name or ''}".strip()
    member = session.query(ChatMember).filter_by(chat_id=chat_id, user_id=user_id).first()
    if not member:
        member = ChatMember(chat_id=chat_id, user_id=user_id, user_name=user_name, full_name=full_name)
        session.add(member)
    else:
        # Обновляем данные, если участник уже существует
        member.full_name = full_name
    session.commit()
    session.close()


def remove_member_from_chat(chat_id, user_id):
    session = Session()
    member = session.query(ChatMember).filter_by(chat_id=chat_id, user_id=user_id).first()
    if member:
        session.delete(member)
        session.commit()
    session.close()


def get_course_start_date(chat_id):
    session = Session()
    chat = session.query(Chat).filter_by(id=chat_id).first()
    if chat:
        return chat.start_date
    session.close()
    return None


def set_course_start_date(chat_id, start_date):
    session = Session()
    chat = session.query(Chat).filter_by(id=chat_id).first()
    if not chat:
        chat = Chat(id=chat_id, start_date=start_date)
        session.add(chat)
    else:
        chat.start_date = start_date
    session.commit()
    session.close()


def set_start_date(update, context):
    chat_id = update.message.chat_id
    user_id = update.message.from_user.id
    bot = context.bot

    # Проверка, является ли пользователь администратором
    if not is_admin(user_id, chat_id, bot):
        update.message.reply_text("Только администраторы могут изменять дату начала курса.")
        return ConversationHandler.END

    try:
        # Извлечение даты из аргументов команды
        start_date_str = context.args[0]
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        set_course_start_date(chat_id, start_date)
        update.message.reply_text(f'Дата начала курса установлена на {start_date}.')
    except (IndexError, ValueError):
        update.message.reply_text('Неверный формат даты. Используйте ГГГГ-ММ-ДД.')

    return ConversationHandler.END


def show_participants(update, context):
    query = update.callback_query
    query.answer()
    chat_id = query.message.chat_id
    session = Session()
    members = session.query(ChatMember).filter_by(chat_id=chat_id).all()
    keyboard = []

    for member in members:
        button_text = f"{member.user_name or member.user_id}"
        callback_data = f"remove_{member.user_id}"
        keyboard.append([InlineKeyboardButton(button_text, callback_data=callback_data)])

    reply_markup = InlineKeyboardMarkup(keyboard)
    query.edit_message_text(text='Выберите участника для удаления:', reply_markup=reply_markup)
    session.close()


def get_all_chats():
    session = Session()  # Предполагается, что Session уже определен в вашем коде
    chats = session.query(Settings).all()
    session.close()
    return chats


def remove_member(update, context):
    chat_id = update.message.chat_id
    user_id = update.message.from_user.id
    bot = context.bot

    if not is_admin(user_id, chat_id, bot):
        update.message.reply_text('Только администраторы могут удалять участников.')
        return

    try:
        user_id_to_remove = int(context.args[0])
        remove_member_from_chat(chat_id, user_id_to_remove)
        update.message.reply_text(f'Участник с ID {user_id_to_remove} удален.')
    except (IndexError, ValueError):
        update.message.reply_text('Неверный формат команды. Используйте /remove <user_id>.')


def handle_new_member(update, context, scheduler, bot):
    for member in update.message.new_chat_members:
        add_member_to_chat(update.message.chat_id, member.id, member.username, member.first_name, member.last_name)
        reschedule_jobs(scheduler, bot)
        check_and_schedule_messages(scheduler, bot)
        check_hashtags_and_notify(bot)


def handle_left_member(update, context):
    remove_member_from_chat(update.message.chat_id, update.message.left_chat_member.id)


def handle_message(update, context):
    if update.edited_message:
        # Обработка отредактированного сообщения
        chat_id = update.edited_message.chat.id
        user_id = update.edited_message.from_user.id
        user_name = update.edited_message.from_user.username
        first_name = update.edited_message.from_user.first_name
        last_name = update.edited_message.from_user.last_name or ""
        new_user_name = update.edited_message.from_user.username
    else:
        # Обработка обычного сообщения
        chat_id = update.message.chat_id
        user_id = update.message.from_user.id
        user_name = update.message.from_user.username
        first_name = update.message.from_user.first_name
        last_name = update.message.from_user.last_name or ""
        new_user_name = update.message.from_user.username

    try:
        chat_member = context.bot.get_chat_member(chat_id, user_id)
        if chat_member.status in ['administrator', 'creator']:
            logger.info(f"User {user_id} in chat {chat_id} is an admin or creator, skipping database addition.")
            return
    except Exception as e:
        logger.error(f"Error checking user status in chat {chat_id} for user {user_id}: {e}")

    session = Session()
    member = session.query(ChatMember).filter_by(chat_id=chat_id, user_id=user_id).first()
    if member:
        # Проверяем, обновился ли user_name
        if member.user_name != new_user_name:
            member.user_name = new_user_name
            session.commit()
    session.close()

    # Получение текста из сообщения или подписи к изображению
    message = update.message if update.message else update.edited_message

    text_to_process = None
    if message:
        if message.text:
            text_to_process = message.text.lower()
        elif message.caption:
            text_to_process = message.caption.lower()

    # Логирование полученных данных
    logger.info(f"Received a message from chat {chat_id}, user {user_id}")
    logger.info(f"Text to process: '{text_to_process}'")

    moscow_tz = pytz.timezone('Europe/Moscow')
    current_time = datetime.now(moscow_tz)
    today_date = current_time.date()
    start_date_str = get_course_start_date(chat_id)

    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        current_time = datetime.now(pytz.timezone('Europe/Moscow'))
        today_date = current_time.date()

        # Расчет номера дня и недели
        day_number = (today_date - start_date).days + 1
        week_number = (day_number - 1) // 7 + 1

        # Проверяем, входит ли текущий день в диапазон 9 недель (63 дня)
        if 1 <= day_number <= 63:
            possible_hashtags = [
                f"#оу{day_number}",
                f"#ов{day_number}",
                f"#неделя{week_number}",
            ]

            if text_to_process:
                logger.info(f"Received a message from chat {chat_id}, user {user_id}")
                logger.info(f"Text to process: '{text_to_process}'")

            else:
                logger.info("The message does not contain text or caption.")

            if any(hashtag in text_to_process for hashtag in possible_hashtags):
                logger.info(f"Hashtag found in text: {text_to_process}")

                add_member_to_chat(chat_id, user_id, user_name, first_name, last_name)
                settings = get_settings(chat_id)

                morning_tag = settings.morning_hashtag if settings.morning_hashtag else "#оу"
                evening_tag = settings.evening_hashtag if settings.evening_hashtag else "#ов"
                week_tag = settings.week_hashtag if settings.week_hashtag else "#неделя"

                morning_deadline = time(10, 1)
                evening_deadline = time(23, 59)

                if f"{morning_tag}{day_number}" in text_to_process and current_time.time() < morning_deadline:
                    update_daily_record(chat_id, user_id, today_date.strftime('%Y-%m-%d'), morning_hashtag=True)
                elif f"{evening_tag}{day_number}" in text_to_process and current_time.time() < evening_deadline:
                    update_daily_record(chat_id, user_id, today_date.strftime('%Y-%m-%d'), evening_hashtag=True)
                elif f"{week_tag}{week_number}" in text_to_process and current_time.weekday() == 6:
                    update_daily_record(chat_id, user_id, today_date.strftime('%Y-%m-%d'), week_hashtag=True)

                if today_date != datetime.now(moscow_tz).date():
                    logger.info("The message is from a previous day, ignoring.")
                    return

                else:
                    logger.info("No fine hashtag found in the text.")
                return
            else:
                logger.info("No relevant hashtag found in the text.")
        else:
            logger.info("The message date is outside the 9-week range.")
    else:
        logger.info("Start date is not set for the chat.")


def send_course_completion_message(bot, chat_id):
    message = (
        "Сердечные поздравления всем участникам нашего захватывающего путешествия в мир знаний! 🌟 Ваша "
        "целеустремлённость и настойчивость поражают воображение, а ваш прогресс вызывает искреннее восхищение. 🚀 Мы "
        "бурно аплодируем вашим успехам и гордимся каждым из вас! Пусть дорога впереди будет освещена светом радости, "
        "благополучия и неустанного стремления к новым вершинам. 🌈🌟 Желаем вам цветущего счастья, ослепительных "
        "успехов и бесконечного финансового процветания. Продолжайте расти и развиваться, и пусть каждый новый шаг "
        "будет наполнен вдохновением и радостью! 🎉 Вы - настоящие герои своей истории, и впереди вас ждут только "
        "самые яркие страницы! 💫 "
    )
    bot.send_message(chat_id=chat_id, text=message)


def schedule_course_completion_message(scheduler, bot, chat_id, start_date_str):
    # Преобразуем строку в объект datetime
    moscow_tz = timezone('Europe/Moscow')
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')

    # Рассчитываем дату завершения курса (62 дня после начала)
    completion_date = start_date + timedelta(days=62)

    # Назначаем время отправки сообщения
    send_time = datetime.combine(completion_date, time(18, 0)).astimezone(pytz.timezone('Europe/Moscow'))

    # Добавляем задачу в планировщик
    scheduler.add_job(send_course_completion_message, 'date', run_date=send_time, args=(bot, chat_id),
                      timezone=moscow_tz)


def check_and_schedule_messages(scheduler, bot):
    session = Session()
    chats = session.query(Chat).all()
    for chat in chats:
        if chat.start_date:
            schedule_course_completion_message(scheduler, bot, chat.id, chat.start_date)
    session.close()


def error(update, context):
    logger.warning('Update "%s" caused error "%s"', update, context.error)


def create_conversation_handler(scheduler, updater):
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("buttons", show_buttons)],
        states={
            'SET_START_DATE': [MessageHandler(Filters.text & ~Filters.command, set_start_date)],
            'REMOVE_MEMBER': [MessageHandler(Filters.text & ~Filters.command, remove_member)],

        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )

    return conv_handler


def test_job():
    print("Test job executed", datetime.now())


def main():
    # Токен вашего бота
    TOKEN = ''

    updater = Updater(TOKEN, use_context=True)
    bot = updater.bot
    # check_hashtags_and_notify(bot)
    dp = updater.dispatcher
    # Настройка планировщика для автоматической проверки хештегов
    scheduler = BackgroundScheduler()
    scheduler.start()
    # morning_check_time = time(10, 1)  # в 10:01 утра
    # evening_check_time = time(23, 59, 59)  # 23:59:59 вечера
    # week_check_time = time(23, 59, 59)
    #
    moscow_tz = pytz.timezone('Europe/Moscow')
    #
    # # Планирование выполнения функции каждый день в 10:00 и 23:59 по Московскому времени
    # job_function = check_hashtags_and_notify(bot)
    # scheduler.add_job(job_function, 'cron', day_of_week='mon,tue,wed,thu,fri,sat', hour=10, minute=1,
    #                   timezone=moscow_tz)
    # scheduler.add_job(job_function, 'cron', day_of_week='mon,tue,wed,thu,fri,sat', hour=23, minute=59, second=59,
    #                   timezone=moscow_tz)
    # scheduler.add_job(job_function, 'cron', day_of_week='sun', hour=23, minute=59, timezone=moscow_tz)
    # scheduler.add_job(lambda: check_reports_and_notify(bot), 'cron', day_of_week='mon,tue,wed,thu,fri,sat',
    #                   hour=morning_check_time.hour,
    #                   minute=morning_check_time.minute, timezone=pytz.timezone('Europe/Moscow'))
    # scheduler.add_job(lambda: check_reports_and_notify(bot), 'cron', day_of_week='mon,tue,wed,thu,fri,sat',
    #                   hour=evening_check_time.hour,
    #                   minute=evening_check_time.minute, second=evening_check_time.second,
    #                   timezone=pytz.timezone('Europe/Moscow'))
    # scheduler.add_job(lambda: check_reports_and_notify(bot), 'cron', day_of_week='sun', hour=week_check_time.hour,
    #                   minute=week_check_time.minute, second=week_check_time.second,
    #                   timezone=pytz.timezone('Europe/Moscow'))

    custom_handle_new_member = partial(handle_new_member, scheduler=scheduler, bot=bot)
    # check_reports_and_notify(bot)
    conv_handler = create_conversation_handler(scheduler, updater)
    reschedule_jobs(scheduler, bot)
    check_and_schedule_messages(scheduler, bot)
    # check_hashtags_and_notify(bot)
    dp.add_handler(conv_handler)
    dp.add_handler(CommandHandler("start", start))
    dp.add_handler(CommandHandler("setstartdate", set_start_date, pass_args=True))
    dp.add_handler(CommandHandler("remove", remove_member, pass_args=True))
    dp.add_handler(CommandHandler('join', join))
    dp.add_handler(CallbackQueryHandler(button))
    dp.add_handler(
        MessageHandler(Filters.update.message & (Filters.text | Filters.caption) & ~Filters.command, handle_message))
    dp.add_handler(MessageHandler(Filters.update.edited_message, handle_message))
    # dp.add_handler(MessageHandler(Filters.status_update.new_chat_members, handle_new_member))
    dp.add_handler(MessageHandler(Filters.status_update.new_chat_members, custom_handle_new_member))
    dp.add_handler(MessageHandler(Filters.status_update.left_chat_member, handle_left_member))
    dp.add_error_handler(error)

    # Запуск бота
    updater.start_polling()
    updater.idle()


if __name__ == '__main__':
    main()
